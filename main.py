from openpyxl import load_workbook


def load_supplier_order(name="supplier_export.xlsx") -> dict:
    """
    Создаем словарь с артикулами и размерами поставщика, которые должны
    были попасть в систему ВБ.
    Расширенная версия нашего заказа, выгружена из внутренней системы.
    У вас может отличаться в каком столбце артикул, размер и баркод, просто
    укажите порядковый номер вашего шаблона, начиная с нуля.
    """
    # Номер колонки в нашем файле из внутренней системы (начиная с нуля)
    supplier_sku_offset = 3
    supplier_size_offset = 4
    supplier_barcode_offset = 6

    wb = load_workbook(name)
    ws = wb.active

    not_found = 0
    found = dict()
    for row in ws.values:
        if not row[0]:
            not_found += 1
            continue
        sku = row[supplier_sku_offset]
        size: str = row[supplier_size_offset]
        barcode = row[supplier_barcode_offset]
        if not size:
            size = "0"
        else:
            size = str(size).replace(".", ",")
            size = size.lower()
        found[sku] = found.get(sku, dict())
        found[sku][size] = barcode

    return found


def load_wb_detalization(order, name="wb_order_export.xlsx") -> dict:
    """
    Это детализация заказа, место, где вб пишет, что есть товары, которые
    он не обработал, но он не скажет нам их баркоды. Так что мучайтесь,
    пишите вот такие скрипты :)
    Берем отсюда https://seller.wb.ru/supply-plan-upload/orders детализация заказа
    """

    # Номер колонки в файле ВБ (начиная с нуля) - как правило, это вам менять не надо
    wb_sku_offset = 2
    wb_size_offset = 3

    wb = load_workbook(name)
    ws = wb.active

    suspicious = dict()
    for row in ws.values:
        if not row[0]:
            continue
        sku = row[wb_sku_offset]
        size: str = row[wb_size_offset]
        size = size.lower()
        try:
            del order[sku][size]
            if not order[sku]:
                del order[sku]
        except KeyError:
            # Иногда размер не совпадает с тем, как у нас во внутренней системе заведено,
            # поэтому создаем отдельный словарь подозрительных артикулов.
            suspicious[sku] = suspicious.get(sku, list()) + [size]

    return suspicious


def print_output(order, suspicious):
    solid_output = []
    possible_output = []
    for key, value in order.items():

        if key in suspicious:
            possible_output.append(
                f"{key} размеры {value} есть у поставщика, у ВБ их нет, но у них есть такие размеры: {suspicious[key]}"
            )
        else:
            solid_output.append(f"{key} размеры {value} не найдены")
    print(f"ТОЧНО НЕ НАЙДЕНЫ:")
    for o in solid_output:
        print(o)
    print(f"НЕ НАЙДЕНЫ, НО У ВБ ЕСТЬ ПОХОЖИЕ РАЗМЕРЫ (ПРОВЕРИТЬ):")
    for o in possible_output:
        print(o)


if __name__ == "__main__":
    order = load_supplier_order("supplier_export.xlsx")
    suspicious = load_wb_detalization(order, "wb_order_export.xlsx")
    print_output(order, suspicious)


